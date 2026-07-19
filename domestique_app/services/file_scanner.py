"""File and image content scanner for sensitive data detection.

Extracts text from various file formats and feeds it through the existing
detection pipeline. Supports:
    - Images (PNG, JPG, TIFF) via Apple Vision OCR or Tesseract fallback
    - PDF documents via pdfminer.six
    - CSV/Excel spreadsheets via csv/openpyxl
    - Plain text / code files (direct read)
    - Base64-encoded content (decoded then routed by type)

Architecture:
    File bytes -> Format detection -> Text extraction -> Detection pipeline -> Result

The scanner is designed for minimal latency:
    - Apple Vision OCR runs on the Neural Engine (hardware-accelerated)
    - Text extraction is synchronous but fast (no network calls)
    - Detection reuses the existing tiered pipeline (regex -> NER -> LLM)
"""

from __future__ import annotations

import base64
import csv
import io
import logging
import time
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Callable

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data types
# ---------------------------------------------------------------------------


class FileType(Enum):
    """Supported file types for scanning."""

    IMAGE = "image"
    PDF = "pdf"
    SPREADSHEET = "spreadsheet"
    CODE = "code"
    TEXT = "text"
    UNKNOWN = "unknown"


@dataclass
class ScanResult:
    """Result of scanning a file for sensitive content."""

    file_name: str
    file_type: FileType
    contains_sensitive: bool
    categories: list[str] = field(default_factory=list)
    detections: list[dict] = field(default_factory=list)
    extracted_text: str = ""
    extraction_time_ms: float = 0.0
    detection_time_ms: float = 0.0
    total_time_ms: float = 0.0
    error: str | None = None


@dataclass
class ExtractionResult:
    """Result of text extraction from a file."""

    text: str
    confidence: float = 1.0
    time_ms: float = 0.0
    method: str = "direct"


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------


def detect_file_type(data: bytes, filename: str = "") -> FileType:
    """Detect file type from magic bytes and/or filename extension.

    Args:
        data: Raw file bytes (at least first 16 bytes needed).
        filename: Optional filename for extension-based fallback.

    Returns:
        Detected FileType enum value.
    """
    if len(data) < 4:
        return FileType.TEXT

    # Magic byte detection
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return FileType.IMAGE
    if data[:2] == b"\xff\xd8":  # JPEG
        return FileType.IMAGE
    if data[:4] == b"II*\x00" or data[:4] == b"MM\x00*":  # TIFF
        return FileType.IMAGE
    if data[:5] == b"%PDF-":
        return FileType.PDF
    if data[:4] == b"PK\x03\x04":  # ZIP-based (xlsx, docx)
        ext = Path(filename).suffix.lower() if filename else ""
        if ext in (".xlsx", ".xls"):
            return FileType.SPREADSHEET
        return FileType.TEXT

    # Extension-based fallback
    ext = Path(filename).suffix.lower() if filename else ""
    if ext in (".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp"):
        return FileType.IMAGE
    if ext == ".pdf":
        return FileType.PDF
    if ext in (".csv", ".xlsx", ".xls", ".tsv"):
        return FileType.SPREADSHEET
    if ext in (
        ".py",
        ".js",
        ".ts",
        ".go",
        ".rs",
        ".java",
        ".rb",
        ".env",
        ".yaml",
        ".yml",
        ".toml",
        ".ini",
        ".cfg",
        ".conf",
        ".sh",
    ):
        return FileType.CODE

    # Handle dotfiles like ".env" where Path(".env").suffix is empty
    name = Path(filename).name.lower() if filename else ""
    if name in (".env", ".envrc", ".netrc", ".npmrc", ".pypirc"):
        return FileType.CODE
    if ext in (".txt", ".md", ".rst", ".log"):
        return FileType.TEXT

    # Try to decode as text
    try:
        data[:1024].decode("utf-8")
        return FileType.TEXT
    except (UnicodeDecodeError, ValueError):
        return FileType.UNKNOWN


# ---------------------------------------------------------------------------
# Text extraction engines
# ---------------------------------------------------------------------------


def extract_text_from_image(data: bytes) -> ExtractionResult:
    """Extract text from image using Apple Vision OCR (preferred) or Tesseract.

    Apple Vision uses the Neural Engine on Apple Silicon for hardware-accelerated
    OCR with excellent accuracy on screenshots and documents.
    """
    start = time.perf_counter()

    # Try Apple Vision first (macOS only, hardware-accelerated)
    try:
        text, confidence = _apple_vision_ocr(data)
        elapsed = (time.perf_counter() - start) * 1000
        return ExtractionResult(
            text=text, confidence=confidence, time_ms=elapsed, method="apple_vision"
        )
    except Exception as e:
        logger.debug(f"Apple Vision OCR failed, trying Tesseract: {e}")

    # Fallback: Tesseract
    try:
        text = _tesseract_ocr(data)
        elapsed = (time.perf_counter() - start) * 1000
        return ExtractionResult(text=text, confidence=0.8, time_ms=elapsed, method="tesseract")
    except Exception as e:
        logger.warning(f"All OCR engines failed: {e}")
        elapsed = (time.perf_counter() - start) * 1000
        return ExtractionResult(text="", confidence=0.0, time_ms=elapsed, method="failed")


def _apple_vision_ocr(data: bytes) -> tuple[str, float]:
    """Perform OCR using Apple's Vision framework (VNRecognizeTextRequest).

    Returns:
        Tuple of (extracted_text, average_confidence).
    """
    import Vision
    from Foundation import NSData

    # Create NSData from bytes
    ns_data = NSData.dataWithBytes_length_(data, len(data))

    # Create image request handler
    handler = Vision.VNImageRequestHandler.alloc().initWithData_options_(ns_data, None)

    # Create text recognition request
    request = Vision.VNRecognizeTextRequest.alloc().init()
    request.setRecognitionLevel_(Vision.VNRequestTextRecognitionLevelAccurate)
    request.setUsesLanguageCorrection_(True)

    # Perform request
    success = handler.performRequests_error_([request], None)
    if not success[0]:
        raise RuntimeError(f"Vision request failed: {success[1]}")

    # Extract results
    results = request.results()
    if not results:
        return "", 0.0

    lines = []
    total_confidence = 0.0
    for observation in results:
        candidate = observation.topCandidates_(1)[0]
        lines.append(candidate.string())
        total_confidence += candidate.confidence()

    text = "\n".join(lines)
    avg_confidence = total_confidence / len(results) if results else 0.0
    return text, avg_confidence


def _tesseract_ocr(data: bytes) -> str:
    """Fallback OCR using Tesseract (pytesseract)."""
    import pytesseract
    from PIL import Image

    img = Image.open(io.BytesIO(data))
    return pytesseract.image_to_string(img)


def extract_text_from_pdf(data: bytes) -> ExtractionResult:
    """Extract text from PDF using pdfminer.six."""
    start = time.perf_counter()
    try:
        from pdfminer.high_level import extract_text as pdf_extract

        text = pdf_extract(io.BytesIO(data))
        elapsed = (time.perf_counter() - start) * 1000
        return ExtractionResult(text=text, time_ms=elapsed, method="pdfminer")
    except Exception as e:
        elapsed = (time.perf_counter() - start) * 1000
        logger.warning(f"PDF extraction failed: {e}")
        return ExtractionResult(text="", confidence=0.0, time_ms=elapsed, method="failed")


def extract_text_from_spreadsheet(data: bytes, filename: str = "") -> ExtractionResult:
    """Extract text from CSV or Excel files."""
    start = time.perf_counter()
    ext = Path(filename).suffix.lower() if filename else ""

    try:
        if ext in (".xlsx", ".xls") or data[:4] == b"PK\x03\x04":
            text = _extract_xlsx(data)
            method = "openpyxl"
        else:
            text = _extract_csv(data)
            method = "csv"

        elapsed = (time.perf_counter() - start) * 1000
        return ExtractionResult(text=text, time_ms=elapsed, method=method)
    except Exception as e:
        elapsed = (time.perf_counter() - start) * 1000
        logger.warning(f"Spreadsheet extraction failed: {e}")
        return ExtractionResult(text="", confidence=0.0, time_ms=elapsed, method="failed")


def _extract_xlsx(data: bytes) -> str:
    """Extract all cell values from Excel workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_csv(data: bytes) -> str:
    """Extract all rows from CSV."""
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        if row:
            lines.append(" | ".join(row))
    return "\n".join(lines)


def extract_text_from_code(data: bytes) -> ExtractionResult:
    """Extract text from code/config files (direct decode)."""
    start = time.perf_counter()
    text = data.decode("utf-8", errors="replace")
    elapsed = (time.perf_counter() - start) * 1000
    return ExtractionResult(text=text, time_ms=elapsed, method="direct")


def extract_text_from_text(data: bytes) -> ExtractionResult:
    """Extract text from plain text files."""
    start = time.perf_counter()
    text = data.decode("utf-8", errors="replace")
    elapsed = (time.perf_counter() - start) * 1000
    return ExtractionResult(text=text, time_ms=elapsed, method="direct")


# ---------------------------------------------------------------------------
# Unified extraction router
# ---------------------------------------------------------------------------


def extract_text(data: bytes, filename: str = "") -> ExtractionResult:
    """Route file to appropriate text extraction engine.

    Args:
        data: Raw file bytes.
        filename: Optional filename for type detection.

    Returns:
        ExtractionResult with extracted text and metadata.
    """
    file_type = detect_file_type(data, filename)

    if file_type == FileType.IMAGE:
        return extract_text_from_image(data)
    elif file_type == FileType.PDF:
        return extract_text_from_pdf(data)
    elif file_type == FileType.SPREADSHEET:
        return extract_text_from_spreadsheet(data, filename)
    elif file_type == FileType.CODE:
        return extract_text_from_code(data)
    elif file_type == FileType.TEXT:
        return extract_text_from_text(data)
    else:
        # Try as text
        try:
            text = data.decode("utf-8", errors="replace")
            return ExtractionResult(text=text, method="fallback_text")
        except Exception:
            return ExtractionResult(text="", confidence=0.0, method="unsupported")


# ---------------------------------------------------------------------------
# Full scan pipeline
# ---------------------------------------------------------------------------


def scan_file(
    data: bytes,
    filename: str = "",
    detector_fn: Callable[[str], list[dict]] | None = None,
) -> ScanResult:
    """Scan a file for sensitive content.

    This is the main entry point. It:
    1. Detects file type
    2. Extracts text content
    3. Runs the detection pipeline on extracted text
    4. Returns structured results

    Args:
        data: Raw file bytes.
        filename: Optional filename for type/extension detection.
        detector_fn: Optional detection function. If None, uses the default
                     regex-based detector from the detection pipeline.
                     Signature: (text: str) -> list[dict] where each dict has
                     at minimum {"category": str, "value": str}.

    Returns:
        ScanResult with detection details and timing.
    """
    total_start = time.perf_counter()
    file_type = detect_file_type(data, filename)

    # Step 1: Extract text
    extraction = extract_text(data, filename)

    if not extraction.text.strip():
        total_ms = (time.perf_counter() - total_start) * 1000
        return ScanResult(
            file_name=filename or "unknown",
            file_type=file_type,
            contains_sensitive=False,
            extraction_time_ms=extraction.time_ms,
            total_time_ms=total_ms,
        )

    # Step 2: Run detection
    detect_start = time.perf_counter()

    if detector_fn is None:
        detector_fn = _default_detector

    detections = detector_fn(extraction.text)
    detect_ms = (time.perf_counter() - detect_start) * 1000

    # Step 3: Build result
    categories = sorted(set(d["category"] for d in detections))
    total_ms = (time.perf_counter() - total_start) * 1000

    return ScanResult(
        file_name=filename or "unknown",
        file_type=file_type,
        contains_sensitive=len(detections) > 0,
        categories=categories,
        detections=detections,
        extracted_text=extraction.text,
        extraction_time_ms=extraction.time_ms,
        detection_time_ms=detect_ms,
        total_time_ms=total_ms,
    )


def scan_base64(
    encoded: str,
    filename: str = "",
    detector_fn: Callable[[str], list[dict]] | None = None,
) -> ScanResult:
    """Scan base64-encoded file content.

    Common in LLM API payloads where images are sent as base64 strings.
    """
    try:
        # Strip data URI prefix if present
        if "," in encoded[:100]:
            encoded = encoded.split(",", 1)[1]
        data = base64.b64decode(encoded)
    except Exception as e:
        return ScanResult(
            file_name=filename or "base64_content",
            file_type=FileType.UNKNOWN,
            contains_sensitive=False,
            error=f"Base64 decode failed: {e}",
        )

    return scan_file(data, filename, detector_fn)


# ---------------------------------------------------------------------------
# Default detector (regex-based, fast)
# ---------------------------------------------------------------------------


def _default_detector(text: str) -> list[dict]:
    """Run regex-based detection on extracted text.

    Uses the same patterns as the main firewall detection pipeline.
    Tuned for high precision (minimize false positives) while maintaining
    good recall on structured PII formats.
    """
    import re

    detections = []
    patterns = {
        "email": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
        "SSN": r"\b\d{3}-\d{2}-\d{4}\b",
        "phone": r"\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}",
        "credit_card": r"\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b",
        "api_key": (
            r"(?:sk-[A-Za-z0-9_-]{20,}|"
            r"AKIA[A-Z0-9]{16,}|"
            r"ghp_[A-Za-z0-9]{30,}|"
            r"xoxb-[A-Za-z0-9-]{20,}|"
            r"sk_live_[A-Za-z0-9]{20,}|"
            r"SG\.[A-Za-z0-9]{20,})"
        ),
        "credential": (
            r"(?:password|passwd|pwd|secret|token|SECRET_ACCESS_KEY|DB_PASSWORD|"
            r"JWT_SECRET|SENDGRID_API_KEY|STRIPE_SECRET|REDIS_URL)"
            r"\s*[=:]\s*['\"]?"
            r"([^\s'\"]{8,})"
        ),
    }

    for category, pattern in patterns.items():
        for match in re.finditer(pattern, text, re.IGNORECASE):
            detections.append(
                {
                    "category": category,
                    "value": match.group(0)[:50],
                    "position": match.start(),
                }
            )

    # Context-aware name detection: look for labeled name fields
    name_patterns = [
        r"(?:Full\s*Name|Name|Customer|Employee|Client)\s*[:=]\s*([A-Z][a-z]+ [A-Z][a-z]+)",
        r"\|\s*([A-Z][a-z]+ [A-Z][a-z]+)\s*\|",  # Table format
    ]
    for pattern in name_patterns:
        for match in re.finditer(pattern, text):
            name = match.group(1)
            # Filter out common false positives
            if name.lower() not in (
                "new york",
                "san francisco",
                "los angeles",
                "main street",
                "team alpha",
            ):
                detections.append(
                    {
                        "category": "name",
                        "value": name,
                        "position": match.start(),
                    }
                )

    # Address detection: look for street patterns with numbers
    addr_pattern = r"\b\d{1,5}\s+[A-Z][a-z]+\s+(?:St|Ave|Dr|Rd|Blvd|Ln|Way|Ct|Pl)\b"
    for match in re.finditer(addr_pattern, text):
        detections.append(
            {
                "category": "address",
                "value": match.group(0),
                "position": match.start(),
            }
        )

    return detections
